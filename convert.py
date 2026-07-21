"""
convert.py
----------
Reads inventory/Inventory.xlsx and generates products.json

Usage:
    python convert.py

Requirements:
    pip install openpyxl
"""

import json
import os
import openpyxl

# ── Config ──
EXCEL_PATH  = os.path.join(os.getcwd(), 'inventory', 'Inventory.xlsx')
OUTPUT_PATH = os.path.join(os.getcwd(), 'products.json')

SINGLE_SHEETS = [
    'Hot Wheels Mainline',
    'Hot Wheels Premium',
    'Hot Wheels Treasure Hunt',
    'Hot Wheels 5 Packs',
    'Matchbox Card',
    'Matchbox Box',
    'Matchbox Moving Parts',
    'MiniGT',
]

BUNDLE_SHEET = 'Bundles'

# Column positions (0-indexed)
COL_CODE           = 0
COL_BRAND          = 1
COL_BUNDLE_NAME    = 2
COL_NAME           = 3
COL_SERIES         = 4
COL_YEAR           = 5
COL_COST_PRICE     = 6
COL_ORIGINAL_PRICE = 7
COL_SELLING_PRICE  = 8
COL_STATUS         = 9


def clean(val):
    if val is None:
        return ''
    return str(val).strip()


def parse_year(val):
    if val is None:
        return None
    try:
        v = int(float(str(val)))
        if v > 40000:
            from datetime import date, timedelta
            excel_epoch = date(1899, 12, 30)
            actual_date = excel_epoch + timedelta(days=v)
            return actual_date.year
        return v
    except:
        return None


def parse_price(val):
    s = clean(val)
    if s.upper() in ('N/A', '', 'NONE', '-'):
        return None
    try:
        return int(float(s))
    except:
        return None


def process_sheet(ws, is_bundle=False):
    products = []
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return products

    for row in rows[1:]:
        if not row or not row[COL_CODE]:
            continue

        code   = clean(row[COL_CODE])
        brand  = clean(row[COL_BRAND])
        name   = clean(row[COL_NAME])
        series = clean(row[COL_SERIES])
        status = clean(row[COL_STATUS]).upper()

        if not code or not name:
            continue

        selling_price = parse_price(row[COL_SELLING_PRICE])

        # Skip only truly invalid entries (no price at all)
        if selling_price is None or selling_price == 0:
            print(f"  ⏭️  Skipping (no price): {code} - {name}")
            continue

        year           = parse_year(row[COL_YEAR])
        bundle_name    = clean(row[COL_BUNDLE_NAME])
        original_price = parse_price(row[COL_ORIGINAL_PRICE])

        # SOLD items are included but marked as SOLD
        item_status = "SOLD" if status == "SOLD" else "Available"

        product = {
            "code":           code,
            "brand":          brand,
            "name":           name,
            "bundle_name":    bundle_name if (is_bundle and bundle_name and bundle_name.upper() != 'N/A') else None,
            "series":         series,
            "year":           year,
            "price":          selling_price,
            "original_price": original_price,
            "status":         item_status,
            "type":           "bundle" if is_bundle else "single",
            "image":          f"images/{code}.jpg"
        }
        products.append(product)

    return products


def main():
    print(f"\n📂 Reading: {EXCEL_PATH}")

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ ERROR: File not found at {EXCEL_PATH}")
        print("   Make sure Inventory.xlsx is in the inventory/ folder.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    all_products = []
    idx = 1
    total_sold = 0
    total_available = 0

    for sheet_name in SINGLE_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet not found, skipping: {sheet_name}")
            continue

        ws = wb[sheet_name]
        products = process_sheet(ws, is_bundle=False)

        if not products:
            print(f"  📭 Empty sheet: {sheet_name}")
            continue

        available = sum(1 for p in products if p['status'] == 'Available')
        sold      = sum(1 for p in products if p['status'] == 'SOLD')
        total_available += available
        total_sold      += sold

        print(f"  ✅ {sheet_name}: {available} available, {sold} sold")

        for p in products:
            p['id'] = idx
            idx += 1
            all_products.append(p)

    if BUNDLE_SHEET in wb.sheetnames:
        ws = wb[BUNDLE_SHEET]
        bundles = process_sheet(ws, is_bundle=True)
        b_available = sum(1 for p in bundles if p['status'] == 'Available')
        b_sold      = sum(1 for p in bundles if p['status'] == 'SOLD')
        print(f"  ✅ {BUNDLE_SHEET}: {b_available} available, {b_sold} sold")
        for p in bundles:
            p['id'] = idx
            idx += 1
            all_products.append(p)
    else:
        print(f"  ⚠️  Sheet not found, skipping: {BUNDLE_SHEET}")

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(all_products, f, indent=2, ensure_ascii=False)

    print(f"\n✅ Done! {len(all_products)} products written to products.json")
    print(f"   Available : {total_available}")
    print(f"   SOLD      : {total_sold}")
    print(f"   Last ID   : {idx - 1}")
    print(f"   Output    : {OUTPUT_PATH}\n")


if __name__ == '__main__':
    main()
